import openpyxl
import numpy as np

files = [
'产品矩阵.md',
'产品配置表_原数据.xlsx',
'常见问题排查_pages 29-32.pdf',
]

workbook = openpyxl.load_workbook(f'resources/{files[1]}')

overrides = {
        'Sheet4': {
            (3,2): 'MOVAP50 Pro',
            (3,3): '科沃斯T50Pro',
            (3,4): '云鲸J5',
            (3,5): '美的V15 Pro',
            (3,6): '追觅S30PU增强版',
            (3,7): '石头P20 Pro'
        },
}

def flat_str(s):
    return str(s).replace(' ', '').replace('\n', '')

def rows_of(s):
    return len([1 for _ in s.iter_rows(values_only=True)])

def cols_of(s):
    return len([1 for _ in s.iter_cols(values_only=True)])

def cell_of(s,row,col):
    return s.cell(row+1, col+1)

def str_cell_of(s,row,col):
    if s.title in overrides and (row, col) in overrides[s.title]:
        return overrides[s.title][(row, col)]
    value = s.cell(row+1, col+1).value

    if s.title == 'Sheet5' and row == 1 and col >= 2:
        value = '追觅' + value

    if s.title == 'Sheet6' and row == 2:
        if col >= 2 and col <= 3:
            value = '科沃斯' + value
        if col >= 4 and col <= 5:
            value = '追觅' + value
        if col >= 6 and col <= 7:
            value = '石头' + value

    if s.title == 'Sheet7' and row == 3 and col >= 2:
        value = '科沃斯' + s.cell(1, col+1).value

    return str(value)

def row_of(s, row, off, length):
    return [flat_str(str_cell_of(s, row, j)) for j in range(off, off+length)]

def col_of(s, col, off, length):
    return [flat_str(str_cell_of(s, j, col)) for j in range(off, off+length)]

def ndarray(rows, cols):
    ret = []
    for _ in range(0, rows):
        ret.append([None] * cols)
    return ret

none_values = {'x', '—', '--.', 'None', '×'}
value_map = {'✔': '支持'}

out = open('resources/out.txt', 'w')

def process(sheet,begin):
    rows, cols = rows_of(sheet), cols_of(sheet)
    begin_row, begin_col = begin

    th = row_of(sheet, begin_row - 1, begin_col, cols - begin_col)

    attrs = col_of(sheet, begin_col - 1, begin_row, rows - begin_row)

    values = ndarray(rows-begin_row, cols-begin_col)

    for i in range(begin_row, rows):
        for j in range(begin_col, cols):
            values[i-begin_row][j-begin_col] = str_cell_of(sheet,i,j)

    for col in range(0, len(th)):
        for row in range(0, len(attrs)):
            if values[row][col] is None or not values[row][col]: continue
            value = flat_str(values[row][col])
            if value in none_values: continue
            if value in value_map: value = value_map[value]
            out.write(f"{th[col]} {attrs[row]}: {value}\n")





process(workbook['Sheet0'], (2,2))
process(workbook['Sheet1'], (2,1))
process(workbook['Sheet2'], (1,2))
process(workbook['Sheet3'], (2,1))
process(workbook['Sheet4'], (4,2))
process(workbook['Sheet5'], (2,2))
process(workbook['Sheet6'], (3,2))
process(workbook['Sheet7'], (4,2))
process(workbook['Sheet8'], (2,2))
process(workbook['Sheet9'], (2,2))

out.close()
